# news.py
import os
import asyncio
import pandas as pd
from fastapi import APIRouter, HTTPException
from pydantic import BaseModel
from openpyxl import load_workbook

router = APIRouter(prefix="/api/news", tags=["News"])

# --- Configuration ---
BASE_DIR = os.path.dirname(os.path.abspath(__file__))
EXCEL_FILE = r"C:\Projects\ml\PulseCI\news_analysis.xlsx"
SHEET_NAME = "news"

# -----------------------------
# Pydantic Models
# -----------------------------
class NewsCreate(BaseModel):
    title: str
    link: str
    source: str
    topic_id: int | None = None

class NewsUpdate(BaseModel):
    title: str | None = None
    link: str | None = None
    source: str | None = None
    topic_id: int | None = None

# -----------------------------
# Helper: Read sheet
# -----------------------------
async def read_news_sheet() -> pd.DataFrame:
    if not os.path.exists(EXCEL_FILE):
        raise HTTPException(status_code=500, detail="Excel database not found.")
    return await asyncio.to_thread(
        pd.read_excel,
        EXCEL_FILE,
        sheet_name=SHEET_NAME
    )

# -----------------------------
# Helper: Save sheet WITHOUT deleting other sheets
# -----------------------------
async def save_news_sheet(df: pd.DataFrame):

    def _write():
        wb = load_workbook(EXCEL_FILE)

        # If sheet exists, delete & rewrite only that sheet
        if SHEET_NAME in wb.sheetnames:
            std = wb[SHEET_NAME]
            wb.remove(std)

        # Create sheet and write manually
        ws = wb.create_sheet(SHEET_NAME)

        # Write header
        ws.append(list(df.columns))

        # Write rows
        for row in df.itertuples(index=False, name=None):
            ws.append(list(row))

        wb.save(EXCEL_FILE)

    await asyncio.to_thread(_write)

# -----------------------------
# GET all news
# -----------------------------
@router.get("/")
async def get_news():
    df = await read_news_sheet()

    # convert NaN → None for JSON
    df = df.where(pd.notnull(df), None)
    df = df.applymap(lambda x: x.item() if hasattr(x, "item") else x)

    return df.to_dict(orient="records")

# -----------------------------
# POST create news
# -----------------------------
@router.post("/")
async def add_news(item: NewsCreate):
    df = await read_news_sheet()

    if df.empty:
        df = pd.DataFrame(columns=["news_id", "title", "link", "source", "topic_id"])

    # Auto-increment news_id
    next_id = int(df["news_id"].max()) + 1 if not df.empty else 1

    new_row = {
        "news_id": next_id,
        "title": item.title,
        "link": item.link,
        "source": item.source,
        "topic_id": item.topic_id,
    }

    df = pd.concat([df, pd.DataFrame([new_row])], ignore_index=True)

    await save_news_sheet(df)

    return {"message": "News item added", "news_id": next_id}

# -----------------------------
# PUT update news
# -----------------------------
@router.put("/{news_id}")
async def update_news(news_id: int, item: NewsUpdate):
    df = await read_news_sheet()

    if news_id not in df["news_id"].values:
        raise HTTPException(status_code=404, detail="News not found")

    # Update only provided fields
    for field, value in item.dict(exclude_none=True).items():
        df.loc[df["news_id"] == news_id, field] = value

    await save_news_sheet(df)
    return {"message": "News updated successfully"}

# -----------------------------
# DELETE news (hard delete)
# -----------------------------
@router.delete("/{news_id}")
async def delete_news(news_id: int):
    df = await read_news_sheet()

    if news_id not in df["news_id"].values:
        raise HTTPException(status_code=404, detail="News not found")

    df = df[df["news_id"] != news_id]

    await save_news_sheet(df)
    return {"message": "News deleted successfully"}
